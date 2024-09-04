import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Load the Excel file and select the 'Saturday' sheet
excel_file_path = r'C:/Users/ishaa/OneDrive/Desktop/assignment1/4bits_01.xlsx'
workbook = openpyxl.load_workbook('C:/Users/ishaa/OneDrive/Desktop/assignment1/4bits_01.xlsx')
sheet_name = 'Saturday'  # Change this if you're working with a different day
sheet = workbook[sheet_name]  # Correctly access the sheet by its name

# Keywords to search
keywords = ["Dhaka", "Baby", "School", "Cricket", "Money", "Int", "Look", "Hello", "By"]

# Set up the WebDriver (Ensure you have ChromeDriver installed)
driver = webdriver.Chrome()

try:
    for i, keyword in enumerate(keywords, start=2):
        # Go to Google
        driver.get('https://www.google.com')
        
        # Find the search box, enter the keyword and press ENTER
        search_box = driver.find_element(By.NAME, 'q')
        search_box.send_keys(keyword)
        search_box.send_keys(Keys.RETURN)
        
        # Wait for the suggestions to load
        WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.sbct .sbl1 span')))

        # Extract the suggestions
        suggestions = driver.find_elements(By.CSS_SELECTOR, '.sbct .sbl1 span')
        texts = [s.text for s in suggestions]

        if texts:
            # Find the longest and shortest suggestions
            longest = max(texts, key=len)
            shortest = min(texts, key=len)
        else:
            longest = ''
            shortest = ''

        # Write the results back to the Excel file
        sheet.cell(row=i, column=2).value = longest  # B column for longest
        sheet.cell(row=i, column=3).value = shortest  # C column for shortest

finally:
    # Close the browser
    driver.quit()

    # Save the modified Excel file
    output_file_path = r'C:/Users/ishaa/OneDrive/Desktop/assignment1/assignment1_updated.xlsx'
    workbook.save(output_file_path)
